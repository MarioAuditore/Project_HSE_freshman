"""Подключаем библиотеку для работы с exсel"""
import openpyxl
"""Открываем файл, указывая путь к нему"""
wb_summary = openpyxl.load_workbook('/Users/alexsoldatov/PycharmProjects/Project/summary.xlsx', data_only=True)
"""Указываем какой лист в докумен будет активным (с какого будем считывать информацию)"""
wb_summary.active = 0
sheet = wb_summary.active
"""Считаем размеры таблицы"""
rows = sheet.max_row  # количество строк
cols = sheet.max_column  # количество столбцов
"""Заполняем массив данными из таблицы, чтобы обратиться к элементу С3
надо обратиться к элементу [2][1]"""
for j in range(1, 11):
    column = [[sheet.cell(row=i, column=j).value for i in range(2, rows + 1)] for j in range(1, cols + 1)]
"""print(column[0][0])"""
campaign_id_summary = [column[0][j] for j in range(0, rows-1)]
"""for j in range(len(campaign_id_summary)):
    print(campaign_id_summary[j])"""
clicks_summary = [column[1][j] for j in range(0, rows-1)]
cv_fact_summary = [column[5][j] for j in range(0, rows-1)]
CPC_fact_summary = [column[7][j] for j in range(0, rows-1)]
CPA_fact_summary = [column[9][j] for j in range(0, rows-1)]
cr_summary = [-1 for j in range(0, len(clicks_summary))]
for j in range(0, len(clicks_summary)-1):
    if (cv_fact_summary[j] != 0) and (clicks_summary[j] != 0):
        cr_summary[j] = (cv_fact_summary[j]/clicks_summary[j])
i = 0
k = len(cr_summary)
while i < k - 1:
    if (cr_summary[i] == -1):
        for j in range(i, k - 1):
            CPC_fact_summary[j] = CPC_fact_summary[j + 1]
            CPA_fact_summary[j] = CPA_fact_summary[j + 1]
            campaign_id_summary[j] = campaign_id_summary[j+1]
            cr_summary[j] = cr_summary[j+1]
            cv_fact_summary[j] = cv_fact_summary[j+1]
            clicks_summary[j] = clicks_summary[j+1]
        del [cr_summary[k-1], cv_fact_summary[k-1], clicks_summary[k-1], campaign_id_summary[k-1], CPC_fact_summary[k-1], CPA_fact_summary[k-1]]
        k -= 1
    else:
        i += 1
if (cr_summary[len(cr_summary)-1] == -1):
    del cr_summary[len(cr_summary)-1]
    del cv_fact_summary[len(cr_summary)-1]
    del clicks_summary[len(clicks_summary) - 1]
    del campaign_id_summary[len(campaign_id_summary) - 1]
    del CPC_fact_summary[len(CPC_fact_summary) - 1]
    del CPA_fact_summary[len(CPA_fact_summary) - 1]

k_coef_1_summary = [-1 for j in range(0, len(campaign_id_summary))]
for j in range(len(campaign_id_summary) - 1):
    k_coef_1_summary[j] = CPC_fact_summary[j]/(CPA_fact_summary[j]*cr_summary[j])
w_coef_2_summary = [-1 for j in range(0, len(campaign_id_summary))]
for j in range(len(campaign_id_summary) - 1):
    a = float(CPA_fact_summary[j])
    b = float(cr_summary[j])
    w_coef_2_summary[j] = clicks_summary[j]/pow((a*b), 0.5)
wb_month_3 = openpyxl.load_workbook('/Users/alexsoldatov/PycharmProjects/Project/avtodom_3month_dates.xlsx', data_only=True)
wb_month_3.active = 0
sheet = wb_month_3.active
rows = sheet.max_row  # количество строк
cols = sheet.max_column  # количество столбцов
for j in range(1, 10):
    column = [[sheet.cell(row=i, column=j).value for i in range(2, rows + 1)] for j in range(1, cols + 1)]
keyword_id_month_3 = [column[0][j] for j in range(0, rows-1)]
campaign_id_month_3 = [column[1][j] for j in range(0, rows-1)]
clicks_month_3 = [column[2][j] for j in range(0, rows-1)]
calls_month_3 = [column[4][j] for j in range(0, rows-1)]
k = len(keyword_id_month_3)
i = 0
while i < k-1:
    if keyword_id_month_3[i] == keyword_id_month_3[i+1]:
        clicks_month_3[i] += clicks_month_3[i+1]
        calls_month_3[i] += calls_month_3[i+1]
        for j in range(i, k-1):
            keyword_id_month_3[j] = keyword_id_month_3[j+1]
            campaign_id_month_3[j] = campaign_id_month_3[j+1]
        del[keyword_id_month_3[k-1], clicks_month_3[k-1], calls_month_3[k-1], campaign_id_month_3[k-1]]
        k -= 1
    else:
        i += 1
if keyword_id_month_3[len(keyword_id_month_3)-1] == keyword_id_month_3[len(keyword_id_month_3)-2]:
    clicks_month_3[len(keyword_id_month_3)-2] += clicks_month_3[len(keyword_id_month_3)-1]
    calls_month_3[len(keyword_id_month_3)-2] += calls_month_3[len(keyword_id_month_3)-1]
    del [keyword_id_month_3[len(keyword_id_month_3)-1], clicks_month_3[len(clicks_month_3)-1]]
    del [calls_month_3[len(calls_month_3)-1], campaign_id_month_3[len(campaign_id_month_3)-1]]

bufer = keyword_id_month_3[0]
flag = 0
cr_finalochka = [0 for i in range(len(keyword_id_month_3))]
for i in range(len(keyword_id_month_3)-1):
    for j in range(len(campaign_id_summary)-1):
        if (campaign_id_summary[j] == campaign_id_month_3[i]):
            bufer = cr_summary[j]
            flag = 1
            cr_finalochka[i] = (calls_month_3[i] + 1)/(clicks_month_3[i] + (1/bufer))
    if (flag == 0):
        cr_finalochka[i] = -2
    flag = 0
r = len(cr_finalochka)
s = 0
while s < r-1:
    if (cr_finalochka[s]==-2) or (cr_finalochka[s] == 0):
        for j in range(s, r-1):
            keyword_id_month_3[j] = keyword_id_month_3[j+1]
            campaign_id_month_3[j] = campaign_id_month_3[j+1]
            cr_finalochka[j] = cr_finalochka[j+1]
        del[keyword_id_month_3[r-1], campaign_id_month_3[r-1], cr_finalochka[r-1]]
        r -= 1
    else:
        s += 1
if cr_finalochka[len(cr_finalochka) - 1] == -2 or cr_finalochka[len(cr_finalochka) - 1] == 0:
    del [keyword_id_month_3[len(cr_finalochka) - 1], campaign_id_month_3[len(cr_finalochka) - 1], cr_finalochka[len(cr_finalochka) - 1]]
"""for j in range(len(cr_finalochka)):
    print(cr_finalochka[j])"""
wb_1month_dates = openpyxl.load_workbook('/Users/alexsoldatov/PycharmProjects/Project/avtodom_1month_dates.xlsx', data_only=True)
"""Указываем какой лист в докумен будет активным (с какого будем считывать информацию)"""
wb_1month_dates.active = 0
sheet = wb_1month_dates.active
"""Считаем размеры таблицы"""
rows = sheet.max_row  # количество строк
cols = sheet.max_column  # количество столбцов
"""Заполняем массив данными из таблицы, чтобы обратиться к элементу С3
надо обратиться к элементу [2][1]"""
for j in range(1, 8):
    column = [[sheet.cell(row=i, column=j).value for i in range(2, rows + 1)] for j in range(1, cols + 1)]
keyword_id_1month_dates = [column[0][j] for j in range(0, rows - 1)]
campaign_id_1month_dates = [column[1][j] for j in range(0, rows - 1)]
bid_average_1month_dates = [column[6][j]/100 for j in range(0, rows -1)]
clicks_word = [-1 for j in range(0, len(keyword_id_1month_dates))]
CPC_word = [-1 for j in range(0, len(keyword_id_1month_dates))]
for j in range(len(campaign_id_1month_dates) - 1):
    for i in range(len(campaign_id_summary) - 1):
        if campaign_id_1month_dates[j] == campaign_id_summary[i]:
            clicks_word[j] = w_coef_2_summary[i]*pow(bid_average_1month_dates[j], 0.5)
            CPC_word[j] = k_coef_1_summary[i]*bid_average_1month_dates[j]
cv_word = [-1 for j in range(0, len(keyword_id_1month_dates))]
for j in range(len(keyword_id_1month_dates) - 1):
    for i in range(len(keyword_id_month_3) - 1):
        if keyword_id_1month_dates[j] == keyword_id_month_3[i]:
            cv_word[j] = cr_finalochka[i]*clicks_word[j]
k = len(campaign_id_1month_dates)
i = 0
while i < k-1:
    if clicks_word[i] == -1:
        for j in range(i, k - 1):
            clicks_word[j] = clicks_word[j+1]
            CPC_word[j] = CPC_word[j+1]
            cv_word[j] = cv_word[j+1]
            campaign_id_1month_dates[j] = campaign_id_1month_dates[j+1]
            keyword_id_1month_dates[j] = keyword_id_1month_dates[j+1]
        del [clicks_word[k - 1], CPC_word[k-1], campaign_id_1month_dates[k-1], keyword_id_1month_dates[k-1], cv_word[k-1]]
        k -= 1
    else:
        i += 1
if clicks_word[len(campaign_id_1month_dates) - 1] == -1:
    del [clicks_word[len(campaign_id_1month_dates) - 1], CPC_word[len(campaign_id_1month_dates) - 1], campaign_id_1month_dates[len(campaign_id_1month_dates) - 1], keyword_id_1month_dates[len(campaign_id_1month_dates) - 1]]
    del [cv_word[len(campaign_id_1month_dates) - 1]]
CPC_count = [1 for j in range(0, len(campaign_id_1month_dates))]
j = 0
k = len(campaign_id_1month_dates)
while j < k - 2:
    i = j + 1
    while i < k - 1:
        if campaign_id_1month_dates[j] == campaign_id_1month_dates[i]:
            clicks_word[j] += clicks_word[i]
            cv_word[j] += cv_word[i]
            CPC_word[j] += CPC_word[i]
            CPC_count[j] += 1
            for t in range(i, k - 2):
                campaign_id_1month_dates[t] = campaign_id_1month_dates[t+1]
                clicks_word[t] = clicks_word[t+1]
                cv_word[t] = cv_word[t+1]
                CPC_word[t] = CPC_word[t+1]
                CPC_count[t] = CPC_count[t+1]
            del [clicks_word[k - 1], campaign_id_1month_dates[k - 1]]
            del [cv_word[k - 1], CPC_word[k - 1]]
            del [CPC_count[k - 1]]
            k -= 1
        else:
            i += 1
    j += 1
for i in range(k - 1):
    CPC_word[i] = CPC_word[i]/CPC_count[i]
"""Запись в таблицу"""
wb_summary = openpyxl.load_workbook('/Users/alexsoldatov/PycharmProjects/Project/summary.xlsx', data_only=True)
"""Указываем какой лист в докумен будет активным (с какого будем считывать информацию)"""
wb_summary.active = 0
sheet = wb_summary.active
"""Считаем размеры таблицы"""
rows = sheet.max_row  # количество строк
cols = sheet.max_column  # количество столбцов
"""Заполняем массив данными из таблицы, чтобы обратиться к элементу С3
надо обратиться к элементу [2][1]"""
for j in range(1, len(campaign_id_1month_dates) + 1):
    campaign_result = sheet.cell(row=j+1, column=13)
    clicks_result = sheet.cell(row=j+1, column=14)
    cpc_result = sheet.cell(row=j+1, column=15)
    cv_result = sheet.cell(row=j+1, column=16)
    campaign_result.value = campaign_id_1month_dates[j-1]
    clicks_result.value = clicks_word[j-1]
    cpc_result.value = CPC_word[j-1]
    cv_result.value = cv_word[j-1]
    wb_summary.save("/Users/alexsoldatov/PycharmProjects/Project/summary.xlsx")
